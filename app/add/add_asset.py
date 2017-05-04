from app import app, db, registry
from app.models import Site, AssetType, FunctionalDescriptor, LoggedEntity, Asset, PointType, AssetPoint, Algorithm
from flask import request, render_template, url_for, redirect, flash, send_file, make_response, jsonify
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.writer.excel import save_virtual_workbook
from io import BytesIO
from xml.etree.ElementTree import ElementTree, Element, SubElement
from base64 import b64encode

# page to add an asset to the site
@app.route('/site/<sitename>/add')
def add_asset_input(sitename):
    site = Site.query.filter_by(name=sitename).one()
    asset_types = AssetType.query.all()
    return render_template('add_asset.html', site=site, asset_types=asset_types)

# return list of functional descriptors through AJAX
@app.route('/site/<sitename>/add/_function')
def return_functions(sitename):
    asset_type = AssetType.query.filter_by(name=request.args['type']).one()
    function_names = [function.name for function in asset_type.functions]
    return jsonify(function_names)

# return list of loggedentities through AJAX
@app.route('/site/<sitename>/add/_loggedentity')
def return_loggedentities(sitename):
    site = Site.query.filter_by(name=sitename).one()

    # get database session for this site
    session = registry.get(site.db_key)

    if not session is None:
        logs = session.query(LoggedEntity).filter_by(type='trend.ETLog').all()
        log_paths = [log.path for log in logs]
        session.close()
    else:
        log_paths = []

    return jsonify(log_paths)

# return list of points through AJAX
@app.route('/site/<sitename>/add/_point')
def return_points(sitename):
    points = PointType.query.all()
    point_names = [point.name for point in points]
    return jsonify(point_names)

# return list of algorithms through AJAX
@app.route('/site/<sitename>/add/_algorithm')
def return_algorithms(sitename):
    algorithms = Algorithm.query.all()
    algorithm_names = [algorithm.descr for algorithm in algorithms]
    return jsonify(algorithm_names)

# process an asset addition
@app.route('/site/<sitename>/add/_submit', methods=['POST'])
def add_asset_submit(sitename):

    # TODO: search by site name will not work if not unique site names
    site = Site.query.filter_by(name=sitename).one()

    # get database session for this site
    session = registry.get(site.db_key)

    # create asset with 0 health
    asset_type = AssetType.query.filter_by(name=request.form['type']).one()
    asset = Asset(type=asset_type, name=request.form['name'], location=request.form['location'], group=request.form['group'], priority=request.form['priority'], notes=request.form['notes'], site=site, health=0)
    db.session.add(asset)
    db.session.commit()

    # TODO: need a better system of reading in values than string-matching point1 and log1
    # generate points
    for i in range(1, len(PointType.query.all()) + 1):
        point_type_name = request.form.get('point' + str(i))
        if not point_type_name is None:

            point_type = PointType.query.filter_by(name=point_type_name).one()
            point = AssetPoint(type=point_type, name=point_type_name)

            # assign the log id to the point
            log_path = request.form.get('log' + str(i))
            if not log_path == '' and not session is None:
                log = session.query(LoggedEntity).filter_by(path=log_path).one()
                point.loggedentity_id = log.id

            asset.points.append(point)

    # set functional descriptors
    function_list = request.form.getlist('function')
    for function_name in function_list:
        function = FunctionalDescriptor.query.filter_by(name=function_name).one()
        asset.functions.append(function)

    # set excluded algorithms
    inclusions = []
    included_list = request.form.getlist('algorithm')
    for algorithm_descr in included_list:
        inclusions.append(Algorithm.query.filter_by(descr=algorithm_descr).one())
    exclusions = set(Algorithm.query.all()) - set(inclusions)
    asset.exclusions.extend(exclusions)

    db.session.commit()
    if not session is None:
        session.close()

    return redirect(url_for('asset_list', sitename=sitename))

# generate a downloadable Excel template for uploading assets
@app.route('/site/<sitename>/add/_download')
def add_asset_download(sitename):

    wb = Workbook()

    # delete initial sheet
    ws_initial = wb.get_sheet_by_name('Sheet')
    wb.remove_sheet(ws_initial)

    for asset_type in AssetType.query.all():
        # generate input sheet for each asset type
        ws_input = wb.create_sheet(asset_type.name)
        ws_input['A1'] = "Name"
        ws_input['B1'] = "Location"
        ws_input['C1'] = "Group"
        ws_input['D1'] = "Priority"
        ws_input['E1'] = "Notes"

        # apply data validation
        priority_dv = DataValidation(type='whole', operator='between', formula1=0, formula2=9)
        ws_input.add_data_validation(priority_dv)
        priority_dv.ranges.append('D2:D1000')

    # save file
    out = BytesIO(save_virtual_workbook(wb))

    # prevent browser from caching the download
    response = make_response(send_file(out, attachment_filename='Template.xlsx', as_attachment=True))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return response

# process an asset addition via upload
@app.route('/site/<sitename>/add/confirm', methods=['POST'])
def add_asset_upload(sitename):

    site = Site.query.filter_by(name=sitename).one()

    # check if file was uploaded
    if 'file' not in request.files:
        flash('No file selected')
        return redirect(url_for('add_asset_input', sitename=sitename))
    file = request.files['file']
    if file.filename =='':
        flash('No file selected')
        return redirect(url_for('add_asset_input', sitename=sitename))

    # check if file is correct type
    if not (file and ('.' in file.filename) and (file.filename.rsplit('.',1)[1].lower() in ['xlsx','xlsm'])):
        flash('File must be xlsx/xlsm')
        return redirect(url_for('add_asset_input', sitename=sitename))

    wb = load_workbook(file)
    asset_list = []

    for ws in wb.worksheets:
        asset_type = AssetType.query.filter_by(name=ws.title).one()

        # generate asset for each non-blank row in the worksheet
        for row in tuple(ws.rows)[1:]:
            # skip if not all required information is present
            if not row[0].value is None and not row[1].value is None and not row[3].value is None:
                name = str(row[0].value)
                location = str(row[1].value)
                if location == "None":
                    location = ""
                group = str(row[2].value)
                if group == "None":
                    group = ""
                priority = row[3].value
                notes = str(row[4].value)
                if notes == "None":
                    notes = ""
                asset = Asset(name=name, location=location, group=group, type=asset_type, priority=priority, notes=notes)
                asset_list.append(asset)

    return render_template('add_asset_confirm.html', assets=asset_list, site=site)

# JSON confirmation for an asset upload
# redirection after form submit occurs from within the HTML form
# this function is only triggered by a javascript AJAX POST request (easiest way to post a JSON object to group data together)
# which is triggered when the submit button is clicked
@app.route('/site/<sitename>/add/_upload_confirm', methods=['POST'])
def add_asset_confirm(sitename):
    site = Site.query.filter_by(name=sitename).one()
    assets_json = request.get_json(force=True)
    # returns a dict with each 'value' being the data for one asset, so iterate through the values
    for asset_json in assets_json.values():
        asset_type = AssetType.query.filter_by(name=asset_json['type']).one()
        asset = Asset(name=asset_json['name'], location=asset_json['location'], group=asset_json['group'], type=asset_type, priority=asset_json['priority'], health=0, site=site)
        db.session.add(asset)
    db.session.commit()
    return True

# page to choose xml downloads
@app.route('/site/<sitename>/assets/download')
def asset_xml_select(sitename):
    site = Site.query.filter_by(name=sitename).one()
    asset_types = AssetType.query.all()
    asset_quantity = {}
    for asset_type in asset_types:
        asset_quantity[asset_type.name] = len(Asset.query.filter_by(site=site, type=asset_type).all())
    return render_template('asset_select.html', assets=site.assets, asset_quantity=asset_quantity, asset_types=asset_types, site=site)

# download xml to import to SBO
@app.route('/site/<sitename>/assets/download/_submit', methods=['POST'])
def asset_xml_download(sitename):

    # select assets
    asset_list = []
    for asset_id in request.form.getlist('asset_id'):
        asset = Asset.query.filter_by(id=asset_id).one()
        asset_list.append(asset)

    # generate xml import to SBO
    object_set = Element('ObjectSet')
    object_set.attrib = {'ExportMode':'Standard', 'Version':'1.8.1.87', 'Note':'TypesFirst'}
    exported_objects = SubElement(object_set, 'ExportedObjects')
    for asset in asset_list:
        oi_folder = SubElement(exported_objects, 'OI')
        oi_folder.attrib = {'NAME':asset.name, 'TYPE':'system.base.Folder'}
        for point in asset.points:
            # generate unique identifier as base64encode(siteID.assetID.pointID)
            # unique id is used for matching medusa points to webreports logs
            identifier = 'DONOTMODIFY:' + str(b64encode('{}.{}.{}'.format(asset.site.id, asset.id, point.id).encode('ascii')).decode('ascii'))

            # create trend log for each point
            oi_trend = SubElement(oi_folder, 'OI')
            oi_trend.attrib = {'NAME':'{} Extended'.format(point.name), 'TYPE':'trend.ETLog', 'DESCR':identifier}

            # flag to include in webreports
            pi = SubElement(oi_trend, 'PI')
            pi.attrib = {'Name':'IncludeInReports', 'Value':'1'}

    # save file
    out = BytesIO()
    tree = ElementTree(object_set)
    tree.write(out, encoding='utf-8', xml_declaration=True)
    out.seek(0)

    # prevent browser from caching the download
    response = make_response(send_file(out, attachment_filename='Import.xml', as_attachment=True))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return response
